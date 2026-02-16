"""Excel 일괄 조회 처리 엔진."""

import io
import json
import os
import re
import tempfile
import threading
import time
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .address import search_vworld, extract_address_components, parse_address, filter_apartment_by_dong_ho

# 전역 작업 관리 (파일 기반, 멀티 워커 호환)
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_JOBS_KEPT = 20
_JOBS_DIR = os.path.join(tempfile.gettempdir(), "siga_batch_jobs")
os.makedirs(_JOBS_DIR, exist_ok=True)

# ─── 업로드 캐시 (파일 기반, 멀티 워커 호환) ───

MAX_UPLOADS_KEPT = 20
_UPLOAD_CACHE_DIR = os.path.join(tempfile.gettempdir(), "siga_upload_cache")
os.makedirs(_UPLOAD_CACHE_DIR, exist_ok=True)

_ADDR_SPLIT_RE = re.compile(r'[\n;]+')
_SHARE_FRAC_RE = re.compile(r'(\d+)\s*/\s*(\d+)')  # 1/2, 3/10
_SHARE_PCT_RE = re.compile(r'([\d.]+)\s*%')  # 50%, 33.33%

# 주소 내 지분 표시 패턴 — "지분" 키워드가 있는 경우만 인식
# 예: "(지분:1/2)", "(지분 1/2)", "(지분1/2)", "지분1/2", "(지분:50%)"
# 주의: 키워드 없는 괄호 분수 (1/2) 등은 본번/부번과 혼동되므로 제외
_ADDR_SHARE_PATTERNS = [
    # 괄호 안: (지분:1/2), (지분 1/2), (지분1/2), (지분:50%), (지분 50%)
    re.compile(r'\(\s*지분[\s:]*(\d+\s*/\s*\d+)\s*\)'),
    re.compile(r'\(\s*지분[\s:]*([\d.]+\s*%)\s*\)'),
    # 괄호 없이: 지분:1/2, 지분 1/2, 지분1/2, 지분:50%, 지분 50%
    re.compile(r'지분[\s:]*(\d+\s*/\s*\d+)'),
    re.compile(r'지분[\s:]*([\d.]+\s*%)'),
]


@dataclass
class UploadCache:
    upload_id: str
    file_bytes: bytes
    headers: list[str]
    row_count: int
    preview: list[list]
    parsed_addresses: list[dict] | None = None
    default_year: str = ""
    created_at: float = field(default_factory=time.time)


@dataclass
class BatchJob:
    job_id: str
    total: int = 0
    processed: int = 0
    current_address: str = ""
    status: str = "pending"  # pending, processing, completed, error
    error: str = ""
    results: list = field(default_factory=list)
    output_bytes: bytes = b""
    created_at: float = field(default_factory=time.time)

    # ─── 파일 기반 상태 관리 ───

    def _job_dir(self) -> str:
        d = os.path.join(_JOBS_DIR, self.job_id)
        os.makedirs(d, exist_ok=True)
        return d

    def save_status(self):
        """현재 상태를 파일에 저장한다 (output_bytes 제외)."""
        data = {
            "job_id": self.job_id,
            "total": self.total,
            "processed": self.processed,
            "current_address": self.current_address,
            "status": self.status,
            "error": self.error,
            "created_at": self.created_at,
        }
        path = os.path.join(self._job_dir(), "status.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)

    def save_output(self):
        """결과 엑셀 바이트를 파일에 저장한다."""
        if self.output_bytes:
            path = os.path.join(self._job_dir(), "output.xlsx")
            with open(path, "wb") as f:
                f.write(self.output_bytes)

    def is_cancelled(self) -> bool:
        """취소 요청이 있는지 확인한다."""
        return os.path.isfile(os.path.join(self._job_dir(), "cancel"))


def cancel_job(job_id: str) -> bool:
    """배치 작업을 취소한다."""
    job_dir = os.path.join(_JOBS_DIR, job_id)
    if not os.path.isdir(job_dir):
        return False
    cancel_path = os.path.join(job_dir, "cancel")
    with open(cancel_path, "w") as f:
        f.write("1")
    return True


def get_job(job_id: str) -> BatchJob | None:
    """파일 기반으로 작업 상태를 로드한다."""
    if not job_id:
        return None
    job_dir = os.path.join(_JOBS_DIR, job_id)
    status_path = os.path.join(job_dir, "status.json")
    if not os.path.isfile(status_path):
        return None
    try:
        with open(status_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        job = BatchJob(
            job_id=data["job_id"],
            total=data.get("total", 0),
            processed=data.get("processed", 0),
            current_address=data.get("current_address", ""),
            status=data.get("status", "pending"),
            error=data.get("error", ""),
            created_at=data.get("created_at", 0),
        )
        # 완료된 경우 output 바이트 로드 (ZIP 우선, 없으면 Excel)
        zip_path = os.path.join(job_dir, "output.zip")
        xlsx_path = os.path.join(job_dir, "output.xlsx")
        if os.path.isfile(zip_path):
            with open(zip_path, "rb") as f:
                job.output_bytes = f.read()
        elif os.path.isfile(xlsx_path):
            with open(xlsx_path, "rb") as f:
                job.output_bytes = f.read()
        return job
    except (json.JSONDecodeError, OSError, KeyError):
        return None


def _cleanup_jobs(keep_id: str = ""):
    """오래된 작업 디렉토리를 정리한다."""
    import shutil
    try:
        entries = []
        for name in os.listdir(_JOBS_DIR):
            d = os.path.join(_JOBS_DIR, name)
            if os.path.isdir(d):
                entries.append((d, os.path.getmtime(d)))
        if len(entries) <= MAX_JOBS_KEPT:
            return
        entries.sort(key=lambda x: x[1])
        for d, _ in entries[: len(entries) - MAX_JOBS_KEPT]:
            if keep_id and os.path.basename(d) == keep_id:
                continue
            shutil.rmtree(d, ignore_errors=True)
    except OSError:
        pass


# ─── 주소 분리 ───

def split_addresses(text: str) -> list[str]:
    """셀 내 복수 주소를 줄바꿈/쉼표/세미콜론으로 분리한다."""
    parts = _ADDR_SPLIT_RE.split(text)
    cleaned = [p.strip() for p in parts if p.strip()]
    return cleaned if cleaned else [text.strip()]


def parse_share(text: str) -> float | None:
    """지분 문자열을 0~1 사이 비율로 변환한다. 실패 시 None 반환."""
    if not text:
        return None
    text = str(text).strip()
    if not text:
        return None

    # 분수: 1/2, 3/10 (분자 < 분모 일 때만 지분으로 인식)
    m = _SHARE_FRAC_RE.search(text)
    if m:
        numer, denom = int(m.group(1)), int(m.group(2))
        if denom > 0 and numer < denom:
            return numer / denom

    # 퍼센트: 50%, 33.33%
    m = _SHARE_PCT_RE.search(text)
    if m:
        val = float(m.group(1)) / 100
        if 0 < val < 1:
            return val

    # 소수: 0.5
    try:
        val = float(text)
        if 0 < val < 1:
            return val
    except (ValueError, TypeError):
        pass

    return None


def format_share(ratio: float | None) -> str:
    """지분 비율을 분수(a/b) 형식으로 변환한다."""
    if ratio is None:
        return ""
    from fractions import Fraction
    frac = Fraction(ratio).limit_denominator(1000)
    if frac.denominator == 1:
        return str(frac.numerator)
    return f"{frac.numerator}/{frac.denominator}"


def extract_share_from_address(addr: str) -> tuple[str, float | None, str]:
    """주소 문자열에서 지분 표시를 추출·제거한다.

    Returns:
        (정리된_주소, 지분_비율, 지분_원본_텍스트(분수부분만))
        지분이 없으면 (원본_주소, None, "")
    """
    for pat in _ADDR_SHARE_PATTERNS:
        m = pat.search(addr)
        if m:
            share_text = m.group(1).strip()
            ratio = parse_share(share_text)
            if ratio is not None:
                # 매치된 전체 부분을 주소에서 제거
                clean = addr[:m.start()] + addr[m.end():]
                clean = clean.strip()
                # 제거 후 남은 불필요한 공백 정리
                clean = re.sub(r'\s{2,}', ' ', clean)
                # 원본 분수/퍼센트 텍스트 그대로 반환
                return clean, ratio, share_text
    return addr, None, ""


# ─── 업로드 캐시 관리 (파일 기반) ───

def _upload_dir(upload_id: str) -> str:
    """업로드 ID에 해당하는 디렉토리 경로를 반환한다."""
    return os.path.join(_UPLOAD_CACHE_DIR, upload_id)


def store_upload(upload_id: str, file_bytes: bytes, headers: list[str],
                 row_count: int, preview: list[list]) -> UploadCache:
    cache = UploadCache(
        upload_id=upload_id,
        file_bytes=file_bytes,
        headers=headers,
        row_count=row_count,
        preview=preview,
    )
    # 파일 기반 저장
    upload_dir = _upload_dir(upload_id)
    os.makedirs(upload_dir, exist_ok=True)

    # 엑셀 원본 바이트
    with open(os.path.join(upload_dir, "file.xlsx"), "wb") as f:
        f.write(file_bytes)

    # 메타데이터 (headers, row_count, preview, created_at)
    meta = {
        "upload_id": upload_id,
        "headers": headers,
        "row_count": row_count,
        "preview": preview,
        "default_year": "",
        "created_at": cache.created_at,
    }
    with open(os.path.join(upload_dir, "meta.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)

    # 오래된 업로드 정리
    _cleanup_uploads(upload_id)

    return cache


def _save_parsed(upload_id: str, parsed: list[dict], default_year: str):
    """파싱된 주소 리스트를 파일에 저장한다."""
    upload_dir = _upload_dir(upload_id)
    if not os.path.isdir(upload_dir):
        return

    data = {"parsed_addresses": parsed, "default_year": default_year}
    with open(os.path.join(upload_dir, "parsed.json"), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def get_upload(upload_id: str) -> UploadCache | None:
    """파일 기반 캐시에서 업로드 데이터를 로드한다."""
    if not upload_id:
        return None
    upload_dir = _upload_dir(upload_id)
    meta_path = os.path.join(upload_dir, "meta.json")
    file_path = os.path.join(upload_dir, "file.xlsx")

    if not os.path.isfile(meta_path) or not os.path.isfile(file_path):
        return None

    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
        with open(file_path, "rb") as f:
            file_bytes = f.read()
    except (json.JSONDecodeError, OSError):
        return None

    cache = UploadCache(
        upload_id=upload_id,
        file_bytes=file_bytes,
        headers=meta.get("headers", []),
        row_count=meta.get("row_count", 0),
        preview=meta.get("preview", []),
        default_year=meta.get("default_year", ""),
        created_at=meta.get("created_at", 0),
    )

    # 파싱 결과가 있으면 로드
    parsed_path = os.path.join(upload_dir, "parsed.json")
    if os.path.isfile(parsed_path):
        try:
            with open(parsed_path, "r", encoding="utf-8") as f:
                parsed_data = json.load(f)
            cache.parsed_addresses = parsed_data.get("parsed_addresses")
            cache.default_year = parsed_data.get("default_year", cache.default_year)
        except (json.JSONDecodeError, OSError):
            pass

    return cache


def _cleanup_uploads(keep_id: str = ""):
    """오래된 업로드 디렉토리를 정리한다."""
    import shutil
    try:
        entries = []
        for name in os.listdir(_UPLOAD_CACHE_DIR):
            d = os.path.join(_UPLOAD_CACHE_DIR, name)
            if os.path.isdir(d):
                entries.append((d, os.path.getmtime(d)))
        if len(entries) <= MAX_UPLOADS_KEPT:
            return
        entries.sort(key=lambda x: x[1])
        for d, _ in entries[: len(entries) - MAX_UPLOADS_KEPT]:
            if keep_id and os.path.basename(d) == keep_id:
                continue
            shutil.rmtree(d, ignore_errors=True)
    except OSError:
        pass


# ─── 병합 셀 처리 ───

def _resolve_merged_cells(ws) -> list[list]:
    """워크시트의 모든 행을 읽되, 병합된 셀의 값을 채운다.

    openpyxl read_only=False 모드에서 사용해야 한다.
    병합된 셀 범위 내의 모든 셀에 상위 좌측 셀의 값을 복사한다.
    """
    # 병합 범위 매핑: (row, col) → 상위 좌측 셀 값
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        top_val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                if (r, c) != (merge_range.min_row, merge_range.min_col):
                    merge_map[(r, c)] = top_val

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = list(row)
        # 병합된 셀 값 채우기
        for c_idx in range(len(vals)):
            key = (row_idx, c_idx + 1)  # 1-based
            if key in merge_map:
                vals[c_idx] = merge_map[key]
        rows.append(vals)

    return rows


# ─── 엑셀 미리보기 ───

def preview_excel(file_bytes: bytes) -> dict:
    """엑셀 파일의 헤더행 + 데이터 행 수 + 미리보기 5행을 반환한다."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    all_rows = _resolve_merged_cells(ws)

    headers = []
    preview_rows = []
    row_count = 0

    for idx, vals in enumerate(all_rows):
        str_vals = [str(c) if c is not None else "" for c in vals]
        if idx == 0:
            headers = str_vals
            continue
        row_count += 1
        if len(preview_rows) < 5:
            preview_rows.append(str_vals)

    wb.close()
    return {
        "headers": headers,
        "row_count": row_count,
        "preview": preview_rows,
    }


# ─── 주소 파싱 ───

def parse_addresses_from_excel(file_bytes: bytes, address_col: int,
                               year_col: int | None,
                               default_year: str,
                               share_col: int | None = None,
                               owner_col: int | None = None) -> list[dict]:
    """선택된 컬럼에서 주소를 추출·분리하고 결과 리스트를 반환한다.

    병합된 셀을 자동으로 해석하여 소유자 등 병합 컬럼 값을 채운다.

    Returns:
        [{row, address, year, share, share_raw, owner, original_row, split_from}, ...]
        split_from 이 있으면 원본 셀에 복수 주소가 있었음을 의미.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # 지분 컬럼 원본 값을 별도로 읽어둔다 (병합 셀 해제 전).
    # openpyxl은 병합 셀의 비-상위좌측 셀에 대해 None을 반환하므로,
    # 병합 해제 없이 읽으면 실제 값이 있는 행만 지분이 표시된다.
    raw_share_by_row: dict[int, str] = {}  # row_idx(0-based) → 원본 지분 텍스트
    if share_col is not None:
        for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
            if row_idx == 0:
                continue  # 헤더 스킵
            if share_col < len(row):
                cell = row[share_col]
                if cell.value is None:
                    continue
                val = cell.value
                if isinstance(val, str):
                    # 텍스트 — 그대로 사용 ("8/50", "50%" 등)
                    text = val.strip()
                    if text:
                        raw_share_by_row[row_idx] = text
                elif isinstance(val, (int, float)):
                    # 숫자 — Excel이 분수를 숫자로 저장한 경우
                    nfmt = cell.number_format or ""
                    fixed_denom = re.search(r'/\s*(\d+)', nfmt)
                    if fixed_denom:
                        # 고정 분모 포맷 (# ??/50 등) → 원본 분수 복원
                        denom = int(fixed_denom.group(1))
                        numer = round(float(val) * denom)
                        raw_share_by_row[row_idx] = f"{numer}/{denom}"
                    elif '/' in nfmt:
                        # 일반 분수 포맷 (# ?/? 등) → Fraction (Excel도 간략화함)
                        from fractions import Fraction
                        frac = Fraction(float(val)).limit_denominator(1000)
                        raw_share_by_row[row_idx] = f"{frac.numerator}/{frac.denominator}"
                    else:
                        # 일반 숫자 — 문자열 그대로
                        text = str(val).strip()
                        if text:
                            raw_share_by_row[row_idx] = text

    all_rows = _resolve_merged_cells(ws)

    results = []
    for idx, vals in enumerate(all_rows):
        if idx == 0:
            continue  # 헤더행 스킵

        if address_col >= len(vals) or not vals[address_col]:
            continue

        raw_addr = str(vals[address_col]).strip()
        if not raw_addr:
            continue

        year = default_year
        if year_col is not None and year_col < len(vals) and vals[year_col]:
            year = str(vals[year_col]).strip()

        # 지분 파싱 — 병합 해제 전 원본 값만 사용 (전이 방지)
        share_raw = raw_share_by_row.get(idx, "")
        share_ratio = parse_share(share_raw) if share_raw else None

        # 소유자
        owner = ""
        if owner_col is not None and owner_col < len(vals) and vals[owner_col]:
            owner = str(vals[owner_col]).strip()

        addresses = split_addresses(raw_addr)
        for addr in addresses:
            # 주소 내 지분 표시 추출 (별도 컬럼 값이 없을 때만 사용)
            clean_addr, embedded_share, embedded_raw = extract_share_from_address(addr)
            if share_ratio is not None:
                # 별도 컬럼 지분이 우선
                final_share = share_ratio
                final_share_raw = share_raw
                final_addr = clean_addr if embedded_share is not None else addr
            elif embedded_share is not None:
                # 주소 내 지분 사용
                final_share = embedded_share
                final_share_raw = embedded_raw
                final_addr = clean_addr
            else:
                final_share = None
                final_share_raw = ""
                final_addr = addr

            results.append({
                "row": idx + 1,  # 엑셀 행 번호 (1-based, 헤더 포함)
                "address": final_addr,
                "year": year,
                "share": final_share,
                "share_raw": final_share_raw,
                "share_display": final_share_raw,
                "owner": owner,
                "original_row": idx + 1,
                "split_from": raw_addr if len(addresses) > 1 else None,
            })

    wb.close()

    return results


# ─── 정리된 주소 엑셀 생성 ───

def generate_cleaned_excel(addresses: list[dict]) -> bytes:
    """파싱된 주소 리스트로 정리된 엑셀을 생성한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "정리된 주소"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    has_share = any(a.get("share") is not None for a in addresses)
    has_owner = any(a.get("owner") for a in addresses)

    headers = ["#"]
    if has_owner:
        headers.append("소유자")
    headers += ["주소", "기준년도"]
    if has_share:
        headers.append("지분")
    headers.append("원본행번호")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 소유자별 정렬 (소유자가 있으면)
    sorted_addrs = addresses
    if has_owner:
        sorted_addrs = sorted(addresses, key=lambda a: (a.get("owner", ""), a.get("original_row", 0)))

    for row_idx, entry in enumerate(sorted_addrs, 2):
        values = [row_idx - 1]
        if has_owner:
            values.append(entry.get("owner", ""))
        values += [
            entry["address"],
            entry["year"],
        ]
        if has_share:
            values.append(entry.get("share_display", ""))
        values.append(entry["original_row"])

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # 열 너비
    for i, h in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        if h == "#":
            ws.column_dimensions[col_letter].width = 5
        elif h == "소유자":
            ws.column_dimensions[col_letter].width = 15
        elif h == "주소":
            ws.column_dimensions[col_letter].width = 40
        elif h in ("기준년도", "지분"):
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 12

    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── 파싱된 주소로 배치 작업 시작 ───

def start_batch_job_from_parsed(
    parsed_addresses: list[dict],
    default_year: str,
    vworld_api_key: str,
    land_module,
    apartment_module,
    house_module,
    building_seoul_module=None,
    flask_app=None,
    pdf_name_pattern: str = "{번호}_{소유자}_{주소}",
) -> str:
    """이미 파싱된 주소 리스트로 배치 작업을 시작하고 job_id를 반환한다."""
    if not parsed_addresses:
        raise ValueError("조회할 주소가 없습니다.")

    rows = [
        {
            "row": a["row"],
            "address": a["address"],
            "year": a["year"] or default_year,
            "share": a.get("share"),
            "share_display": a.get("share_display", ""),
            "owner": a.get("owner", ""),
        }
        for a in parsed_addresses
    ]

    job_id = uuid.uuid4().hex[:12]
    job = BatchJob(job_id=job_id, total=len(rows))
    job.save_status()
    _cleanup_jobs(job_id)

    t = threading.Thread(
        target=_process_batch,
        args=(job, rows, default_year, vworld_api_key,
              land_module, apartment_module, house_module,
              building_seoul_module,
              flask_app, pdf_name_pattern),
        daemon=True,
    )
    t.start()
    return job_id


# ─── 배치 처리 ───

_UNSAFE_FNAME_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _make_pdf_filename(pattern: str, idx: int, owner: str, address: str,
                       year: str, share_display: str) -> str:
    """PDF 파일명을 패턴에 따라 생성한다."""
    # 주소 축약 (시구동 + 건물명 정도)
    addr_short = address
    if len(addr_short) > 40:
        addr_short = addr_short[:40]
    name = pattern.replace("{번호}", str(idx))
    name = name.replace("{소유자}", owner or "")
    name = name.replace("{주소}", addr_short)
    name = name.replace("{년도}", year or "")
    name = name.replace("{지분}", share_display or "")
    # 파일명 정리
    name = _UNSAFE_FNAME_RE.sub("", name).strip()
    name = re.sub(r'_+', '_', name).strip('_')
    if not name:
        name = f"{idx}"
    return name + ".pdf"


def _process_batch(
    job: BatchJob,
    rows: list[dict],
    default_year: str,
    vworld_api_key: str,
    land_module,
    apartment_module,
    house_module,
    building_seoul_module=None,
    flask_app=None,
    pdf_name_pattern: str = "{번호}_{소유자}_{주소}",
):
    """백그라운드에서 각 주소를 조회하고 개별 PDF를 생성한다."""
    job.status = "processing"
    job.save_status()

    pdf_dir = os.path.join(job._job_dir(), "pdfs")
    os.makedirs(pdf_dir, exist_ok=True)
    pdf_files = []  # (filename, filepath) 리스트

    for idx, entry in enumerate(rows, 1):
        # 취소 확인
        if job.is_cancelled():
            job.status = "cancelled"
            job.error = f"사용자에 의해 중지됨 ({job.processed}/{job.total}건 처리 완료)"
            break

        addr_str = entry["address"]
        year = entry["year"] or default_year
        job.current_address = addr_str

        share = entry.get("share")
        share_display = entry.get("share_display", "")
        owner = entry.get("owner", "")

        result_row = {
            "row_num": entry["row"],
            "owner": owner,
            "address": addr_str,
            "pnu": "",
            "year": year,
            "share": share,
            "share_display": share_display,
            "land_price_per_sqm": "",
            "land_area": "",
            "land_total_price": "",
            "land_total_price_share": "",
            "apt_building_name": "",
            "apt_price": "",
            "apt_price_share": "",
            "house_price": "",
            "house_price_share": "",
            "building_name": "",
            "building_total": "",
            "building_total_share": "",
            "status": "",
        }

        # PDF용 섹션 데이터
        auto_sections = {}

        try:
            # 1. 주소 검색 → PNU 추출
            address = _resolve_address_for_batch(addr_str, vworld_api_key)
            pnu = address.get("pnu", "")
            result_row["pnu"] = pnu

            if not pnu and not address.get("adm_cd", ""):
                result_row["status"] = "PNU 추출 실패"
                job.results.append(result_row)
                job.processed += 1
                job.save_status()
                time.sleep(0.3)
                continue

            found_types = []

            # 2. 토지 조회
            try:
                land_result = land_module.search(address, year)
                if land_result.success and land_result.results:
                    first = land_result.results[0]
                    result_row["land_price_per_sqm"] = first.get("price_per_sqm", "")
                    result_row["land_area"] = first.get("land_area", "")
                    result_row["land_total_price"] = first.get("total_price", "")
                    if share is not None and first.get("total_price"):
                        try:
                            result_row["land_total_price_share"] = int(float(first["total_price"]) * share)
                        except (ValueError, TypeError):
                            pass
                    found_types.append("토지")
                    auto_sections["land"] = {
                        "label": "토지 공시지가",
                        "result": {"results": land_result.results},
                        "source_key": "land",
                    }
            except Exception:
                pass
            time.sleep(0.5)

            # 3. 공동주택 조회
            try:
                apt_result = apartment_module.search(address, year)
                if apt_result.success and apt_result.results:
                    # 동/호 필터링 적용
                    dong_no = address.get("dong_no", "")
                    ho_no = address.get("ho_no", "")
                    if dong_no or ho_no:
                        apt_result.results = filter_apartment_by_dong_ho(
                            apt_result.results, dong_no, ho_no)

                if apt_result.success and apt_result.results:
                    first = apt_result.results[0]
                    result_row["apt_building_name"] = first.get("building_name", "")
                    result_row["apt_price"] = first.get("price", "")
                    if share is not None and first.get("price"):
                        try:
                            result_row["apt_price_share"] = int(float(first["price"]) * share)
                        except (ValueError, TypeError):
                            pass
                    found_types.append("공동주택")
                    auto_sections["apartment"] = {
                        "label": "공동주택 공시가격",
                        "result": {"results": apt_result.results},
                        "source_key": "apartment",
                    }
            except Exception:
                pass
            time.sleep(0.5)

            # 4. 개별주택 조회
            try:
                house_result = house_module.search(address, year)
                if house_result.success and house_result.results:
                    first = house_result.results[0]
                    result_row["house_price"] = first.get("price", "")
                    if share is not None and first.get("price"):
                        try:
                            result_row["house_price_share"] = int(float(first["price"]) * share)
                        except (ValueError, TypeError):
                            pass
                    found_types.append("개별주택")
                    auto_sections["house"] = {
                        "label": "개별주택 공시가격",
                        "result": {"results": house_result.results},
                        "source_key": "house",
                    }
            except Exception:
                pass

            # 5. 주택외건물 조회 (서울 ETAX)
            sido = address.get("sido", "")
            if building_seoul_module and sido in ("서울", "서울특별시"):
                try:
                    bldg_result = building_seoul_module.search(address, year)
                    if bldg_result.success and bldg_result.results:
                        # ETAX는 서버 측 dong/ho 필터링을 하지만, 클라이언트 측 보완 필터링
                        dong_no = address.get("dong_no", "")
                        ho_no = address.get("ho_no", "")
                        if dong_no or ho_no:
                            bldg_result.results = filter_apartment_by_dong_ho(
                                bldg_result.results, dong_no, ho_no)

                    if bldg_result.success and bldg_result.results:
                        first = bldg_result.results[0]
                        result_row["building_name"] = first.get("name", "")
                        result_row["building_total"] = first.get("total", "")
                        if share is not None and first.get("total"):
                            try:
                                price_str = re.sub(r'[^\d]', '', str(first["total"]))
                                if price_str:
                                    result_row["building_total_share"] = int(int(price_str) * share)
                            except (ValueError, TypeError):
                                pass
                        found_types.append("주택외건물")
                        auto_sections["building"] = {
                            "label": "주택외건물 시가표준액 (서울)",
                            "result": {"results": bldg_result.results},
                            "source_key": "building_etax",
                        }
                except Exception:
                    pass
                time.sleep(0.5)

            # 공동주택 또는 건물이 있으면 토지 제거 (auto 검색과 동일 로직)
            if "apartment" in auto_sections or "building" in auto_sections:
                auto_sections.pop("land", None)

            if found_types:
                result_row["status"] = ", ".join(found_types)
            else:
                result_row["status"] = "결과없음"

            # PDF 생성
            if auto_sections and flask_app:
                try:
                    pdf_bytes = _generate_single_pdf(
                        flask_app, addr_str, year, auto_sections,
                        owner=owner, share_display=share_display,
                    )
                    fname = _make_pdf_filename(
                        pdf_name_pattern, idx, owner, addr_str,
                        year, share_display,
                    )
                    fpath = os.path.join(pdf_dir, fname)
                    with open(fpath, "wb") as f:
                        f.write(pdf_bytes)
                    pdf_files.append((fname, fpath))
                except Exception:
                    pass  # PDF 실패해도 계속 진행

        except Exception as e:
            result_row["status"] = f"오류: {e}"

        job.results.append(result_row)
        job.processed += 1
        job.save_status()
        time.sleep(0.3)

    # ZIP 생성 (PDF + 요약 엑셀)
    try:
        excel_bytes = _generate_output_excel(job.results)

        # 요약 엑셀 별도 저장 (엑셀만 다운로드용)
        xlsx_path = os.path.join(job._job_dir(), "output.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(excel_bytes)

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, fpath in pdf_files:
                zf.write(fpath, fname)
            zf.writestr("조회결과_요약.xlsx", excel_bytes)
        job.output_bytes = zip_buf.getvalue()

        # ZIP 파일을 상태보다 먼저 저장 (race condition 방지:
        # 브라우저가 cancelled/completed 상태를 보는 시점에 output.zip이 이미 존재해야 함)
        zip_path = os.path.join(job._job_dir(), "output.zip")
        with open(zip_path, "wb") as f:
            f.write(job.output_bytes)

        if job.status != "cancelled":
            job.status = "completed"
        job.save_status()
    except Exception as e:
        job.status = "error"
        job.error = f"ZIP 생성 오류: {e}"
        job.save_status()


def _generate_single_pdf(flask_app, address: str, year: str,
                         sections: dict, owner: str = "",
                         share_display: str = "") -> bytes:
    """개별 부동산의 조회 결과를 PDF 바이트로 생성한다."""
    from weasyprint import HTML
    from .pdf import _safe_url_fetcher
    from .base import SOURCE_INFO

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # auto_data 구조 생성
    auto_data = {
        "address": address,
        "year": year,
        "sections": sections,
        "owner": owner,
        "share_display": share_display,
    }

    with flask_app.app_context():
        from flask import render_template
        html_string = render_template(
            "pdf/auto_pdf.html",
            auto_data=auto_data,
            now=now,
            source_info=SOURCE_INFO,
        )

    pdf_bytes = HTML(
        string=html_string,
        url_fetcher=_safe_url_fetcher,
    ).write_pdf()
    return pdf_bytes


def _resolve_address_for_batch(addr_str: str, vworld_api_key: str) -> dict:
    """배치용 주소 해석: V-World 검색 → 첫 번째 결과에서 PNU 추출."""
    search_result = search_vworld(addr_str, vworld_api_key)
    items = search_result.get("items", [])

    if items:
        addr = extract_address_components(items[0])
        addr["_raw"] = addr_str
    else:
        # V-World 검색 실패 시 파싱으로 폴백
        addr = parse_address(addr_str)
        addr["_raw"] = addr_str

    # parse_address로 dong_no, ho_no 추출하여 병합
    parsed = parse_address(addr_str)
    if parsed.get("dong_no") and not addr.get("dong_no"):
        addr["dong_no"] = parsed["dong_no"]
    if parsed.get("ho_no") and not addr.get("ho_no"):
        addr["ho_no"] = parsed["ho_no"]

    return addr


def _generate_output_excel(results: list[dict]) -> bytes:
    """결과 리스트로부터 엑셀 파일 바이트를 생성한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일괄조회 결과"

    # 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    price_font = Font(bold=True, color="2563EB")

    # 지분/소유자/건물 데이터가 있는지 확인
    has_share = any(r.get("share") is not None for r in results)
    has_owner = any(r.get("owner") for r in results)
    has_building = any(r.get("building_name") or r.get("building_total") for r in results)

    # 소유자가 있으면 소유자별 정렬
    if has_owner:
        results = sorted(results, key=lambda r: (r.get("owner", ""), r.get("row_num", 0)))

    headers = ["#"]
    if has_owner:
        headers.append("소유자")
    headers += ["주소", "PNU", "기준년도"]
    if has_share:
        headers.append("지분")
    headers += [
        "토지_공시지가(원/m²)", "토지_면적(m²)", "토지_총액(원)",
    ]
    if has_share:
        headers.append("토지_지분적용(원)")
    headers += [
        "공동주택_건물명", "공동주택_공시가격(원)",
    ]
    if has_share:
        headers.append("공동주택_지분적용(원)")
    headers += [
        "개별주택_공시가격(원)",
    ]
    if has_share:
        headers.append("개별주택_지분적용(원)")
    if has_building:
        headers += ["주택외건물_소재지", "주택외건물_시가표준액(원)"]
        if has_share:
            headers.append("주택외건물_지분적용(원)")
    headers.append("조회상태")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    share_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    share_font = Font(bold=True, color="D97706")

    for row_idx, row_data in enumerate(results, 2):
        values = [row_idx - 1]
        if has_owner:
            values.append(row_data.get("owner", ""))
        values += [
            row_data.get("address", ""),
            row_data.get("pnu", ""),
            row_data.get("year", ""),
        ]
        if has_share:
            values.append(row_data.get("share_display", ""))
        values += [
            _to_number(row_data.get("land_price_per_sqm", "")),
            _to_number(row_data.get("land_area", "")),
            _to_number(row_data.get("land_total_price", "")),
        ]
        if has_share:
            values.append(_to_number(row_data.get("land_total_price_share", "")))
        values += [
            row_data.get("apt_building_name", ""),
            _to_number(row_data.get("apt_price", "")),
        ]
        if has_share:
            values.append(_to_number(row_data.get("apt_price_share", "")))
        values += [
            _to_number(row_data.get("house_price", "")),
        ]
        if has_share:
            values.append(_to_number(row_data.get("house_price_share", "")))
        if has_building:
            values += [
                row_data.get("building_name", ""),
                _to_number(row_data.get("building_total", "")),
            ]
            if has_share:
                values.append(_to_number(row_data.get("building_total_share", "")))
        values.append(row_data.get("status", ""))

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

        # 가격 컬럼 + 지분적용 컬럼 스타일
        for col_idx, val in enumerate(values, 1):
            header_name = headers[col_idx - 1] if col_idx <= len(headers) else ""
            cell = ws.cell(row=row_idx, column=col_idx)
            if "지분적용" in header_name and val:
                cell.font = share_font
                cell.fill = share_fill
                cell.number_format = "#,##0"
            elif header_name in ("토지_공시지가(원/m²)", "토지_총액(원)",
                                 "공동주택_공시가격(원)", "개별주택_공시가격(원)",
                                 "주택외건물_시가표준액(원)") and val:
                cell.font = price_font
                cell.number_format = "#,##0"

    # 열 너비 설정
    for i in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        h = headers[i - 1]
        if h == "#":
            ws.column_dimensions[col_letter].width = 5
        elif h == "주소":
            ws.column_dimensions[col_letter].width = 35
        elif h == "PNU":
            ws.column_dimensions[col_letter].width = 22
        elif h == "소유자":
            ws.column_dimensions[col_letter].width = 15
        elif h in ("기준년도", "지분"):
            ws.column_dimensions[col_letter].width = 10
        elif "건물명" in h or "소재지" in h:
            ws.column_dimensions[col_letter].width = 20
        elif "면적" in h:
            ws.column_dimensions[col_letter].width = 12
        elif "조회상태" in h:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 18

    # 필터 설정
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_number(val):
    """문자열 숫자를 int/float로 변환, 실패 시 원본 반환."""
    if not val:
        return ""
    try:
        f = float(val)
        if f == int(f):
            return int(f)
        return f
    except (ValueError, TypeError):
        return val


def generate_sample_template() -> bytes:
    """샘플 양식 엑셀 파일을 생성한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일괄조회 양식"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["주소 (필수)", "기준년도 (선택)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    examples = [
        ("서울 강남구 역삼동 601", "2025"),
        ("부산 해운대구 우동 456", ""),
        ("수원시 팔달구 인계동 789", "2024"),
    ]
    for row_idx, (addr, yr) in enumerate(examples, 2):
        ws.cell(row=row_idx, column=1, value=addr)
        ws.cell(row=row_idx, column=2, value=yr)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    # 안내 행
    note_row = len(examples) + 3
    note_cell = ws.cell(row=note_row, column=1, value="[안내] A열에 주소를 입력하세요. B열 기준년도는 선택사항입니다.")
    note_cell.font = Font(color="94A3B8", italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
